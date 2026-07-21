"""
Admin/Dashboard routes
Estadísticas, ranking, choferes online
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from uuid import UUID
from typing import Optional
from datetime import datetime, timedelta, date
from fastapi.responses import Response

from app.database import get_db
from app.dependencies import get_current_admin_user, get_current_user
from app.schemas.admin_schemas import (
    EstadisticasResponse,
    RankingChoferResponse,
    SolicitudActivaResponse
)

router = APIRouter(prefix="/api/control-base", tags=["Dashboard"])


@router.get("/datos")
async def obtener_datos_empresa(
    current_user: tuple = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Get company/fleet information
    """
    control_base_id = current_user[1]
    
    query = text("""
        SELECT id, nombre, email, telefono, latitud, longitud, activo, created_at, updated_at
        FROM tenant.control_base
        WHERE id = :control_base_id
    """)
    
    result = await db.execute(query, {"control_base_id": control_base_id})
    row = result.first()
    
    if not row:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    
    return {
        "id": row[0],
        "nombre": row[1],
        "email": row[2],
        "telefono": row[3],
        "latitud": row[4],
        "longitud": row[5],
        "activo": row[6],
        "created_at": row[7],
        "updated_at": row[8]
    }


@router.put("/actualizar")
async def actualizar_empresa(
    nombre: str = None,
    email: str = None,
    telefono: str = None,
    current_user: tuple = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Update company information
    """
    control_base_id = current_user[1]
    
    updates = []
    params = {"control_base_id": control_base_id}
    
    if nombre:
        updates.append("nombre = :nombre")
        params["nombre"] = nombre
    if email:
        updates.append("email = :email")
        params["email"] = email
    if telefono:
        updates.append("telefono = :telefono")
        params["telefono"] = telefono
    
    if not updates:
        return {"success": True, "message": "No hay datos para actualizar"}
    
    updates.append("updated_at = NOW()")
    
    query = text(f"""
        UPDATE tenant.control_base
        SET {', '.join(updates)}
        WHERE id = :control_base_id
    """)
    
    await db.execute(query, params)
    await db.commit()
    
    return {"success": True, "message": "Datos actualizados"}


# ============================================================
# CHOFERES ONLINE - PARA ADMIN (con filtro de conexión)
# ============================================================

@router.get("/choferes-online")
async def listar_choferes_online(
    current_user: tuple = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Get online drivers (for map display) - SOLO ADMIN
    """
    control_base_id = current_user[1]
    
    query = text("""
        SELECT 
            cv.id,
            u.id as usuario_id,
            COALESCE(p.nombre || ' ' || p.apellido, u.email) as nombre,
            cv.latitud,
            cv.longitud,
            cv.estado_laboral,
            cv.calificacion_promedio,
            v.patente,
            v.modelo,
            cv.total_viajes
        FROM fleet.chofer_vehiculo cv
        JOIN auth.usuario u ON u.id = cv.usuario_id
        LEFT JOIN auth.perfil_general p ON p.usuario_id = u.id
        LEFT JOIN fleet.vehiculo v ON v.id = cv.vehiculo_id
        WHERE cv.control_base_id = :control_base_id
          AND cv.estado_laboral IN ('libre', 'ocupado')
          AND cv.activo = true
          AND cv.ultima_conexion > NOW() - INTERVAL '5 minutes'
        ORDER BY cv.estado_laboral, cv.calificacion_promedio DESC
    """)
    
    result = await db.execute(query, {"control_base_id": control_base_id})
    rows = result.all()
    
    return [
        {
            "id": row[0],
            "usuario_id": row[1],
            "nombre": row[2],
            "latitud": float(row[3]) if row[3] else None,
            "longitud": float(row[4]) if row[4] else None,
            "estado": row[5],
            "calificacion": float(row[6]) if row[6] else 5.0,
            "patente": row[7],
            "modelo": row[8],
            "total_viajes": row[9] or 0
        }
        for row in rows
    ]


# ============================================================
# CHOFERES DISPONIBLES - PARA EMPLEADOS (SOLO LECTURA)
# ============================================================

@router.get("/choferes-disponibles")
async def choferes_disponibles_empleado(
    control_base_id: UUID = Query(..., description="ID de la base operativa"),
    current_user: tuple = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Obtiene choferes disponibles para EMPLEADOS (solo visualización).
    No requiere permisos de administrador.
    """
    query = text("""
        SELECT 
            cv.id,
            u.id as usuario_id,
            COALESCE(p.nombre || ' ' || p.apellido, u.email) as nombre,
            cv.estado_laboral,
            v.patente,
            v.marca,
            v.modelo,
            cv.latitud,
            cv.longitud,
            cv.calificacion_promedio,
            cv.total_viajes
        FROM fleet.chofer_vehiculo cv
        JOIN auth.usuario u ON u.id = cv.usuario_id
        LEFT JOIN auth.perfil_general p ON p.usuario_id = u.id
        LEFT JOIN fleet.vehiculo v ON v.id = cv.vehiculo_id
        WHERE cv.control_base_id = :control_base_id
          AND cv.estado_laboral = 'libre'
          AND cv.activo = true
          AND u.activo = true
        ORDER BY p.nombre ASC
    """)
    
    result = await db.execute(query, {"control_base_id": control_base_id})
    rows = result.all()
    
    return [
        {
            "id": str(row[0]),
            "usuario_id": str(row[1]),
            "nombre": row[2] or "Sin nombre",
            "estado_laboral": row[3],
            "vehiculo": {
                "patente": row[4],
                "marca": row[5] or "Sin marca",
                "modelo": row[6] or "Sin modelo"
            },
            "ubicacion": {
                "latitud": float(row[7]) if row[7] else None,
                "longitud": float(row[8]) if row[8] else None
            },
            "calificacion": float(row[9]) if row[9] else 0,
            "total_viajes": row[10] or 0
        }
        for row in rows
    ]


# ============================================================
# ALIAS PARA COMPATIBILIDAD CON EL FRONTEND (OPCIONAL)
# ============================================================

@router.get("/lista-choferes-online")
async def lista_choferes_online_alias(
    control_base_id: UUID = Query(..., description="ID de la base operativa"),
    current_user: tuple = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Alias de /choferes-online para compatibilidad con el frontend.
    """
    return await listar_choferes_online(current_user, db)


# ============================================================
# SOLICITUDES ACTIVAS
# ============================================================

@router.get("/solicitudes-activas", response_model=list[SolicitudActivaResponse])
async def listar_solicitudes_activas(
    limit: int = 50,
    current_user: tuple = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Get active trip requests (pending and accepted)
    """
    control_base_id = current_user[1]
    
    query = text("""
        SELECT 
            vs.id,
            COALESCE(p.nombre || ' ' || p.apellido, u.email) as pasajero_nombre,
            vs.direccion_origen as origen,
            vs.direccion_destino as destino,
            vs.created_at as solicitado_en,
            EXTRACT(EPOCH FROM (NOW() - vs.created_at))::int as tiempo_espera,
            vs.estado
        FROM trip.viaje_solicitado vs
        JOIN auth.usuario u ON u.id = vs.pasajero_id
        LEFT JOIN auth.perfil_general p ON p.usuario_id = u.id
        WHERE vs.control_base_id = :control_base_id
          AND vs.estado IN ('pendiente', 'aceptado')
        ORDER BY vs.created_at ASC
        LIMIT :limit
    """)
    
    result = await db.execute(query, {
        "control_base_id": control_base_id,
        "limit": limit
    })
    rows = result.all()
    
    return [
        SolicitudActivaResponse(
            viaje_id=row[0],
            pasajero_nombre=row[1],
            origen=row[2],
            destino=row[3],
            solicitado_en=row[4],
            tiempo_espera_segundos=row[5]
        )
        for row in rows
    ]


# ============================================================
# ESTADÍSTICAS
# ============================================================

@router.get("/estadisticas", response_model=EstadisticasResponse)
async def obtener_estadisticas(
    current_user: tuple = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Get dashboard KPIs and metrics
    """
    control_base_id = current_user[1]
    
    query = text("""
        SELECT 
            COUNT(CASE WHEN DATE(vs.created_at) = CURRENT_DATE THEN 1 END) as viajes_hoy,
            COUNT(CASE WHEN DATE_TRUNC('month', vs.created_at) = DATE_TRUNC('month', CURRENT_DATE) THEN 1 END) as viajes_mes,
            COUNT(CASE WHEN cv.ultima_conexion > NOW() - INTERVAL '5 minutes' THEN 1 END) as choferes_online,
            COUNT(DISTINCT cv.id) as choferes_totales,
            COUNT(DISTINCT CASE WHEN tu.nombre = 'pasajero' THEN u.id END) as pasajeros_totales,
            COALESCE(SUM(CASE WHEN DATE(vs.finalizado_en) = CURRENT_DATE THEN vs.precio_final END), 0) as ingresos_hoy,
            COALESCE(SUM(CASE WHEN DATE_TRUNC('month', vs.finalizado_en) = DATE_TRUNC('month', CURRENT_DATE) THEN vs.precio_final END), 0) as ingresos_mes,
            COALESCE(AVG(c.puntaje)::DECIMAL(3,2), 5.0) as calificacion_promedio
        FROM tenant.control_base cb
        LEFT JOIN fleet.chofer_vehiculo cv ON cv.control_base_id = cb.id
        LEFT JOIN auth.usuario u ON u.control_base_id = cb.id
        LEFT JOIN auth.tipo_usuario tu ON u.tipo_usuario_id = tu.id
        LEFT JOIN trip.viaje_solicitado vs ON vs.control_base_id = cb.id
        LEFT JOIN trip.calificacion c ON c.viaje_id = vs.id
        WHERE cb.id = :control_base_id
        GROUP BY cb.id
    """)
    
    result = await db.execute(query, {"control_base_id": control_base_id})
    row = result.first()
    
    if not row:
        return EstadisticasResponse(
            total_viajes_hoy=0,
            total_viajes_mes=0,
            choferes_online=0,
            choferes_totales=0,
            pasajeros_totales=0,
            ingresos_hoy=0.0,
            ingresos_mes=0.0,
            calificacion_promedio=5.0
        )
    
    return EstadisticasResponse(
        total_viajes_hoy=row[0] or 0,
        total_viajes_mes=row[1] or 0,
        choferes_online=row[2] or 0,
        choferes_totales=row[3] or 0,
        pasajeros_totales=row[4] or 0,
        ingresos_hoy=float(row[5] or 0),
        ingresos_mes=float(row[6] or 0),
        calificacion_promedio=float(row[7] or 5.0)
    )


@router.get("/ranking-choferes", response_model=list[RankingChoferResponse])
async def ranking_choferes(
    limit: int = 10,
    criterio: str = "calificacion",
    current_user: tuple = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Get top rated drivers ranking with multiple criteria
    criterio: calificacion, viajes, antiguedad
    """
    control_base_id = current_user[1]
    
    order_by = {
        "calificacion": "cv.calificacion_promedio DESC, cv.total_viajes DESC",
        "viajes": "cv.total_viajes DESC, cv.calificacion_promedio DESC",
        "antiguedad": "cv.created_at ASC"
    }.get(criterio, "cv.calificacion_promedio DESC, cv.total_viajes DESC")
    
    query = text(f"""
        SELECT 
            u.id as chofer_id,
            COALESCE(p.nombre || ' ' || p.apellido, u.email) as nombre,
            cv.calificacion_promedio,
            cv.total_calificaciones,
            cv.total_viajes,
            cv.created_at as fecha_registro,
            p.foto_perfil_url,
            v.patente,
            v.modelo,
            ROW_NUMBER() OVER (ORDER BY {order_by}) as posicion
        FROM fleet.chofer_vehiculo cv
        JOIN auth.usuario u ON u.id = cv.usuario_id
        LEFT JOIN auth.perfil_general p ON p.usuario_id = u.id
        LEFT JOIN fleet.vehiculo v ON v.id = cv.vehiculo_id
        WHERE cv.control_base_id = :control_base_id
          AND cv.total_calificaciones > 0
          AND cv.activo = true
        ORDER BY {order_by}
        LIMIT :limit
    """)
    
    result = await db.execute(query, {
        "control_base_id": control_base_id,
        "limit": limit
    })
    rows = result.all()
    
    return [
        RankingChoferResponse(
            chofer_id=row[0],
            nombre=row[1],
            calificacion_promedio=float(row[2] or 5.0),
            total_viajes=row[4] or 0,
            imagen_url=row[6]
        )
        for row in rows
    ]


@router.get("/choferes-pendientes")
async def listar_choferes_pendientes(
    current_user: tuple = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """
    List pending driver approvals
    """
    control_base_id = current_user[1]
    
    query = text("""
        SELECT 
            cv.id as chofer_vehiculo_id,
            u.id as usuario_id,
            u.email,
            COALESCE(p.nombre, '') as nombre,
            COALESCE(p.apellido, '') as apellido,
            p.telefono,
            v.patente,
            v.marca,
            v.modelo,
            v.anio,
            cv.created_at as registro_fecha
        FROM fleet.chofer_vehiculo cv
        JOIN auth.usuario u ON u.id = cv.usuario_id
        LEFT JOIN auth.perfil_general p ON p.usuario_id = u.id
        JOIN fleet.vehiculo v ON v.id = cv.vehiculo_id
        WHERE cv.control_base_id = :control_base_id
          AND cv.estado_aprobacion = 'pendiente'
        ORDER BY cv.created_at ASC
    """)
    
    result = await db.execute(query, {"control_base_id": control_base_id})
    rows = result.all()
    
    return [
        {
            "chofer_vehiculo_id": str(row[0]),
            "usuario_id": str(row[1]),
            "email": row[2],
            "nombre": row[3],
            "apellido": row[4],
            "telefono": row[5],
            "vehiculo": {
                "patente": row[6],
                "marca": row[7],
                "modelo": row[8],
                "anio": row[9]
            },
            "fecha_registro": row[10]
        }
        for row in rows
    ]


@router.put("/aprobar-chofer/{chofer_vehiculo_id}")
async def aprobar_chofer(
    chofer_vehiculo_id: UUID,
    current_user: tuple = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Approve a pending driver
    """
    control_base_id = current_user[1]
    
    check_query = text("""
        SELECT id FROM fleet.chofer_vehiculo 
        WHERE id = :id AND control_base_id = :control_base_id AND estado_aprobacion = 'pendiente'
    """)
    result = await db.execute(check_query, {
        "id": chofer_vehiculo_id,
        "control_base_id": control_base_id
    })
    
    if not result.first():
        raise HTTPException(status_code=404, detail="Chofer pendiente no encontrado")
    
    update_query = text("""
        UPDATE fleet.chofer_vehiculo
        SET estado_aprobacion = 'aprobado', updated_at = NOW()
        WHERE id = :id
    """)
    
    await db.execute(update_query, {"id": chofer_vehiculo_id})
    await db.commit()
    
    return {"success": True, "message": "Chofer aprobado correctamente"}


@router.put("/rechazar-chofer/{chofer_vehiculo_id}")
async def rechazar_chofer(
    chofer_vehiculo_id: UUID,
    current_user: tuple = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Reject a pending driver
    """
    control_base_id = current_user[1]
    
    check_query = text("""
        SELECT id FROM fleet.chofer_vehiculo 
        WHERE id = :id AND control_base_id = :control_base_id AND estado_aprobacion = 'pendiente'
    """)
    result = await db.execute(check_query, {
        "id": chofer_vehiculo_id,
        "control_base_id": control_base_id
    })
    
    if not result.first():
        raise HTTPException(status_code=404, detail="Chofer pendiente no encontrado")
    
    update_query = text("""
        UPDATE fleet.chofer_vehiculo
        SET estado_aprobacion = 'rechazado', activo = false, updated_at = NOW()
        WHERE id = :id
    """)
    
    await db.execute(update_query, {"id": chofer_vehiculo_id})
    await db.commit()
    
    return {"success": True, "message": "Chofer rechazado"}


# ============================================
# ESTADÍSTICAS AVANZADAS
# ============================================

@router.get("/estadisticas-avanzadas")
async def obtener_estadisticas_avanzadas(
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    current_user: tuple = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Obtener estadísticas avanzadas con gráficos y métricas
    """
    control_base_id = current_user[1]
    
    if not fecha_desde:
        fecha_desde = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not fecha_hasta:
        fecha_hasta = datetime.now().strftime('%Y-%m-%d')
    
    try:
        fecha_desde_date = date.fromisoformat(fecha_desde)
        fecha_hasta_date = date.fromisoformat(fecha_hasta)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    
    params = {
        "control_base_id": control_base_id,
        "fecha_desde": fecha_desde_date,
        "fecha_hasta": fecha_hasta_date
    }
    
    # 1. Resumen General
    query_resumen = text("""
        SELECT 
            COUNT(vs.id) as total_viajes,
            COUNT(CASE WHEN vs.estado = 'finalizado' THEN 1 END) as viajes_completados,
            COUNT(CASE WHEN vs.estado = 'cancelado' THEN 1 END) as viajes_cancelados,
            COUNT(CASE WHEN vs.estado IN ('pendiente', 'aceptado', 'en_curso') THEN 1 END) as viajes_activos,
            COALESCE(SUM(vs.precio_final), 0) as ingresos_totales,
            COALESCE(AVG(vs.precio_final), 0) as promedio_viaje,
            COUNT(DISTINCT vs.pasajero_id) as pasajeros_unicos,
            COUNT(DISTINCT vs.chofer_id) as choferes_activos,
            COALESCE(AVG(c.puntaje), 0) as calificacion_promedio
        FROM trip.viaje_solicitado vs
        LEFT JOIN trip.calificacion c ON c.viaje_id = vs.id
        WHERE vs.control_base_id = :control_base_id
          AND DATE(vs.created_at) BETWEEN :fecha_desde AND :fecha_hasta
    """)
    
    result = await db.execute(query_resumen, params)
    row = result.first()
    
    resumen = {
        "total_viajes": row[0] or 0,
        "viajes_completados": row[1] or 0,
        "viajes_cancelados": row[2] or 0,
        "viajes_activos": row[3] or 0,
        "ingresos_totales": float(row[4] or 0),
        "promedio_viaje": float(row[5] or 0),
        "pasajeros_unicos": row[6] or 0,
        "choferes_activos": row[7] or 0,
        "calificacion_promedio": float(row[8] or 0)
    }
    
    # 2. Viajes por día (tendencia)
    query_tendencia = text("""
        SELECT 
            DATE(vs.created_at) as fecha,
            COUNT(vs.id) as total_viajes,
            COUNT(CASE WHEN vs.estado = 'finalizado' THEN 1 END) as completados,
            COUNT(CASE WHEN vs.estado = 'cancelado' THEN 1 END) as cancelados,
            COALESCE(SUM(vs.precio_final), 0) as ingresos
        FROM trip.viaje_solicitado vs
        WHERE vs.control_base_id = :control_base_id
          AND DATE(vs.created_at) BETWEEN :fecha_desde AND :fecha_hasta
        GROUP BY DATE(vs.created_at)
        ORDER BY fecha ASC
    """)
    
    result = await db.execute(query_tendencia, params)
    rows = result.all()
    
    tendencia = [
        {
            "fecha": row[0].isoformat() if row[0] else None,
            "total_viajes": row[1] or 0,
            "completados": row[2] or 0,
            "cancelados": row[3] or 0,
            "ingresos": float(row[4] or 0)
        }
        for row in rows
    ]
    
    # 3. Distribución de estados
    query_estados = text("""
        SELECT 
            vs.estado,
            COUNT(vs.id) as cantidad
        FROM trip.viaje_solicitado vs
        WHERE vs.control_base_id = :control_base_id
          AND DATE(vs.created_at) BETWEEN :fecha_desde AND :fecha_hasta
        GROUP BY vs.estado
    """)
    
    result = await db.execute(query_estados, params)
    rows = result.all()
    
    estados = {row[0]: row[1] or 0 for row in rows}
    
    # 4. Distribución de calificaciones
    query_calificaciones = text("""
        SELECT 
            c.puntaje,
            COUNT(c.id) as cantidad
        FROM trip.calificacion c
        JOIN trip.viaje_solicitado vs ON vs.id = c.viaje_id
        WHERE vs.control_base_id = :control_base_id
          AND DATE(vs.created_at) BETWEEN :fecha_desde AND :fecha_hasta
        GROUP BY c.puntaje
        ORDER BY c.puntaje ASC
    """)
    
    result = await db.execute(query_calificaciones, params)
    rows = result.all()
    
    calificaciones = {row[0]: row[1] or 0 for row in rows}
    
    # 5. Top Comercios
    query_top_comercios = text("""
        SELECT 
            c.nombre,
            c.rubro,
            COUNT(vs.id) as total_viajes,
            COALESCE(SUM(vs.precio_final), 0) as ingresos
        FROM trip.viaje_solicitado vs
        JOIN public.comercio c ON c.id = vs.comercio_id
        WHERE vs.control_base_id = :control_base_id
          AND vs.comercio_id IS NOT NULL
          AND DATE(vs.created_at) BETWEEN :fecha_desde AND :fecha_hasta
        GROUP BY c.id, c.nombre, c.rubro
        ORDER BY total_viajes DESC
        LIMIT 10
    """)
    
    result = await db.execute(query_top_comercios, params)
    rows = result.all()
    
    top_comercios = [
        {
            "nombre": row[0],
            "rubro": row[1],
            "total_viajes": row[2] or 0,
            "ingresos": float(row[3] or 0)
        }
        for row in rows
    ]
    
    # 6. Top Empresas
    query_top_empresas = text("""
        SELECT 
            e.nombre,
            e.tipo,
            COUNT(vs.id) as total_viajes,
            COALESCE(SUM(vs.precio_final), 0) as ingresos
        FROM trip.viaje_solicitado vs
        JOIN tenant.empresa e ON e.id = vs.empresa_id
        WHERE vs.control_base_id = :control_base_id
          AND vs.empresa_id IS NOT NULL
          AND DATE(vs.created_at) BETWEEN :fecha_desde AND :fecha_hasta
        GROUP BY e.id, e.nombre, e.tipo
        ORDER BY total_viajes DESC
        LIMIT 10
    """)
    
    result = await db.execute(query_top_empresas, params)
    rows = result.all()
    
    top_empresas = [
        {
            "nombre": row[0],
            "tipo": row[1],
            "total_viajes": row[2] or 0,
            "ingresos": float(row[3] or 0)
        }
        for row in rows
    ]
    
    # 7. Ingresos por mes (últimos 12 meses)
    query_ingresos_mensuales = text("""
        SELECT 
            TO_CHAR(DATE_TRUNC('month', vs.created_at), 'YYYY-MM') as mes,
            DATE_TRUNC('month', vs.created_at) as fecha_mes,
            COUNT(vs.id) as total_viajes,
            COALESCE(SUM(vs.precio_final), 0) as ingresos
        FROM trip.viaje_solicitado vs
        WHERE vs.control_base_id = :control_base_id
          AND vs.created_at > NOW() - INTERVAL '12 months'
        GROUP BY DATE_TRUNC('month', vs.created_at)
        ORDER BY fecha_mes ASC
    """)
    
    result = await db.execute(query_ingresos_mensuales, {"control_base_id": control_base_id})
    rows = result.all()
    
    ingresos_mensuales = [
        {
            "mes": row[0],
            "fecha": row[1].isoformat() if row[1] else None,
            "total_viajes": row[2] or 0,
            "ingresos": float(row[3] or 0)
        }
        for row in rows
    ]
    
    return {
        "resumen": resumen,
        "tendencia": tendencia,
        "estados": estados,
        "calificaciones": calificaciones,
        "top_comercios": top_comercios,
        "top_empresas": top_empresas,
        "ingresos_mensuales": ingresos_mensuales,
        "periodo": {
            "desde": fecha_desde,
            "hasta": fecha_hasta
        }
    }


@router.get("/estadisticas-avanzadas/exportar")
async def exportar_estadisticas(
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    formato: str = "excel",
    current_user: tuple = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Exportar estadísticas a Excel o CSV
    """
    stats = await obtener_estadisticas_avanzadas(fecha_desde, fecha_hasta, current_user, db)
    
    if formato == "csv":
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        writer.writerow(["RESUMEN GENERAL"])
        writer.writerow(["Métrica", "Valor"])
        for key, value in stats["resumen"].items():
            writer.writerow([key.replace("_", " ").title(), value])
        
        writer.writerow([])
        writer.writerow(["TENDENCIA DIARIA"])
        writer.writerow(["Fecha", "Total Viajes", "Completados", "Cancelados", "Ingresos"])
        for item in stats["tendencia"]:
            writer.writerow([
                item["fecha"],
                item["total_viajes"],
                item["completados"],
                item["cancelados"],
                item["ingresos"]
            ])
        
        writer.writerow([])
        writer.writerow(["TOP COMERCIOS"])
        writer.writerow(["Nombre", "Rubro", "Total Viajes", "Ingresos"])
        for item in stats["top_comercios"]:
            writer.writerow([
                item["nombre"],
                item["rubro"],
                item["total_viajes"],
                item["ingresos"]
            ])
        
        writer.writerow([])
        writer.writerow(["TOP EMPRESAS"])
        writer.writerow(["Nombre", "Tipo", "Total Viajes", "Ingresos"])
        for item in stats["top_empresas"]:
            writer.writerow([
                item["nombre"],
                item["tipo"],
                item["total_viajes"],
                item["ingresos"]
            ])
        
        csv_content = output.getvalue()
        output.close()
        
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=estadisticas_{datetime.now().strftime('%Y%m%d')}.csv"
            }
        )
    
    try:
        from openpyxl import Workbook
        from io import BytesIO
        
        wb = Workbook()
        
        ws1 = wb.active
        ws1.title = "Resumen"
        ws1.append(["Métrica", "Valor"])
        for key, value in stats["resumen"].items():
            ws1.append([key.replace("_", " ").title(), value])
        
        ws2 = wb.create_sheet("Tendencia Diaria")
        ws2.append(["Fecha", "Total Viajes", "Completados", "Cancelados", "Ingresos"])
        for item in stats["tendencia"]:
            ws2.append([
                item["fecha"],
                item["total_viajes"],
                item["completados"],
                item["cancelados"],
                item["ingresos"]
            ])
        
        ws3 = wb.create_sheet("Top Comercios")
        ws3.append(["Nombre", "Rubro", "Total Viajes", "Ingresos"])
        for item in stats["top_comercios"]:
            ws3.append([
                item["nombre"],
                item["rubro"],
                item["total_viajes"],
                item["ingresos"]
            ])
        
        ws4 = wb.create_sheet("Top Empresas")
        ws4.append(["Nombre", "Tipo", "Total Viajes", "Ingresos"])
        for item in stats["top_empresas"]:
            ws4.append([
                item["nombre"],
                item["tipo"],
                item["total_viajes"],
                item["ingresos"]
            ])
        
        ws5 = wb.create_sheet("Ingresos Mensuales")
        ws5.append(["Mes", "Total Viajes", "Ingresos"])
        for item in stats["ingresos_mensuales"]:
            ws5.append([
                item["mes"],
                item["total_viajes"],
                item["ingresos"]
            ])
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=estadisticas_{datetime.now().strftime('%Y%m%d')}.xlsx"
            }
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")